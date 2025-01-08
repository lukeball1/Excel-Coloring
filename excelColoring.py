import openpyxl
from openpyxl.styles import PatternFill

# Load the spreadsheet
#workbook = openpyxl.load_workbook("your_file.xlsx")
#spreadsheet = workbook.active

#set global color fills
greenFill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
redFill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


#with open("FileDescriptorOrPath", "r") as f:
    # do stuff with file here
    #content = f.readLine()
    # find the muni.name and store that in a string variable
content = 'table1<-table1%>%add_row(candi.name="David A. Becker", num_votes=53,muni.name="winchester al dakarte",type="Alderman",ward="1",year-1993,station="QUE 28")'
muniStartIndex = content.find("muni.name=\"")
muniEndIndex = content.find("\"", muniStartIndex + 12)
muniName = content[muniStartIndex + 11:muniEndIndex]
print(muniName)
    # find the year and store that as a string by getting the next 4 indexes
    
    # As long as year.isdigit() is true and muni.name exists, color the sqare green